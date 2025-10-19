# core/reporting_excel.py
# Version i ri me formatim të avancuar për raportet e analizës së telefonatave
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from pathlib import Path

def write_excel_report_textual(rows, out_path):
    """
    Krijon një raport Excel me dy nivele:
    1. Faqja 'Përmbledhje' që liston çdo agjent dhe përshkrimin e tij të përgjithshëm.
    2. Një faqe për secilin agjent me seksionet 'preggi' dhe 'da migliorare'.
    rows: list[dict] me fushat ['agent','summary','preggi','da_migliorare']
    """
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Përmbledhje"

    # Stil bazë
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Kolonat për faqen kryesore
    ws_summary.append(["Agjenti", "Përmbledhje"])
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 120

    # Grupi sipas agjentit
    by_agent = {}
    for r in rows:
        ag = (r.get("agent") or "UNKNOWN").strip()
        if ag not in by_agent:
            by_agent[ag] = {
                "summary": r.get("summary", ""),
                "preggi": r.get("preggi", ""),
                "da_migliorare": r.get("da_migliorare", "")
            }
        # shto në faqen Përmbledhje vetëm një herë për agjent
        if not any(ws_summary.cell(row=i, column=1).value == ag for i in range(2, ws_summary.max_row+1)):
            ws_summary.append([ag, r.get("summary", "")])

    # Fletët individuale për agjentët
    for agent in sorted(by_agent.keys()):
        data = by_agent[agent]
        ws = wb.create_sheet(title=agent[:30])  # max 31 karaktere në titullin e faqes
        ws.column_dimensions["A"].width = 140

        # pregji
        ws["A1"] = "preggi"
        ws["A1"].font = bold
        ws["A2"] = data["preggi"]
        ws["A2"].alignment = wrap

        # da migliorare
        ws["A4"] = "da migliorare"
        ws["A4"].font = bold
        ws["A5"] = data["da_migliorare"]
        ws["A5"].alignment = wrap

    # Ruaj rezultatin
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def write_excel_report_telemarketing_format(rows, out_path, language="sq"):
    """
    Krijon një raport Excel në formatin e ri me:
    1. Faqja kryesore me header, logo dhe tabelën e agjentëve (vetëm summary)
    2. Faqet individuale për agjentët me strengths dhe improvements
    """
    wb = Workbook()
    
    # Përkthimet bazuar në gjuhën
    translations = {
        "sq": {
            "title": "TRAJNIM TELEMARKETING",
            "nr": "Nr",
            "emer": "Emer", 
            "vleresimi": "Vleresimi",
            "shenime": "Shenime",
            "strengths": "Pikat e Forta",
            "improvements": "Për tu Përmirësuar"
        },
        "it": {
            "title": "FORMAZIONE TELEMARKETING",
            "nr": "Nr",
            "emer": "Nome",
            "vleresimi": "Valutazione", 
            "shenime": "Note",
            "strengths": "Punti di Forza",
            "improvements": "Da Migliorare"
        },
        "en": {
            "title": "TELEMARKETING TRAINING",
            "nr": "No",
            "emer": "Name",
            "vleresimi": "Score",
            "shenime": "Notes",
            "strengths": "Strengths",
            "improvements": "Improvements"
        }
    }
    
    t = translations.get(language, translations["sq"])
    
    # Fshi faqen default
    wb.remove(wb.active)
    
    # Krijo faqen kryesore me emrin e duhur bazuar në gjuhën
    ws_main = wb.create_sheet(t["title"], 0)
    
    # Stilizime të avancuara
    bold = Font(bold=True, size=12)
    title_font = Font(bold=True, size=18, color="1F4E79")
    subtitle_font = Font(bold=True, size=14, color="2F4F4F")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")
    yellow_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    light_green_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    thick_border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )
    
    # Header me logo (nëse ekziston)
    logo_path = Path("assets/protrade.jpg")
    if logo_path.exists():
        try:
            img = Image(logo_path)
            img.width = 200
            img.height = 60
            ws_main.add_image(img, "B1")
        except Exception:
            pass  # Nëse logo nuk mund të ngarkohet, vazhdo pa të
    
    # Titulli kryesor
    title_cell = ws_main["B5"]
    title_cell.value = t["title"]
    title_cell.font = title_font
    title_cell.alignment = center
    
    # Grupi sipas agjentit
    by_agent = {}
    for r in rows:
        ag = (r.get("agent") or "UNKNOWN").strip()
        if ag not in by_agent:
            by_agent[ag] = {
                "summary": r.get("summary", ""),
                "preggi": r.get("preggi", ""),
                "da_migliorare": r.get("da_migliorare", ""),
                "score": r.get("score", 0.0)
            }
    
    # Tabela e agjentëve
    row_start = 8
    ws_main.cell(row=row_start, column=2, value=t["nr"]).font = bold
    ws_main.cell(row=row_start, column=3, value=t["emer"]).font = bold
    ws_main.cell(row=row_start, column=4, value=t["vleresimi"]).font = bold
    ws_main.cell(row=row_start, column=5, value=t["shenime"]).font = bold
    
    # Stilizo header-in e tabelës
    for col in range(2, 6):
        cell = ws_main.cell(row=row_start, column=col)
        cell.fill = yellow_fill
        cell.border = thick_border
        cell.alignment = center
    
    # Shto të dhënat e agjentëve
    for idx, (agent, data) in enumerate(sorted(by_agent.items()), 1):
        row = row_start + idx
        
        # Nr
        ws_main.cell(row=row, column=2, value=idx).border = black_border
        ws_main.cell(row=row, column=2).alignment = center
        
        # Emer (emri i plotë i agjentit)
        ws_main.cell(row=row, column=3, value=agent).border = black_border
        ws_main.cell(row=row, column=3).alignment = center
        
        # Vleresimi (pikësimi numerik)
        score_cell = ws_main.cell(row=row, column=4, value=data["score"])
        score_cell.border = black_border
        score_cell.alignment = center
        # Ngjyrosje bazuar në pikësimin
        if data["score"] >= 4.0:
            score_cell.fill = light_green_fill
        elif data["score"] >= 3.0:
            score_cell.fill = light_blue_fill
        
        # Shenime (vetëm summary, pa titull)
        ws_main.cell(row=row, column=5, value=data["summary"]).border = black_border
        ws_main.cell(row=row, column=5).alignment = wrap
    
    # Rregullo gjerësinë e kolonave
    ws_main.column_dimensions["A"].width = 5
    ws_main.column_dimensions["B"].width = 8
    ws_main.column_dimensions["C"].width = 25
    ws_main.column_dimensions["D"].width = 15
    ws_main.column_dimensions["E"].width = 60
    
    # Faqet individuale për agjentët me strengths dhe improvements
    for agent in sorted(by_agent.keys()):
        data = by_agent[agent]
        ws = wb.create_sheet(title=agent[:30])  # max 31 karaktere në titullin e faqes
        ws.column_dimensions["A"].width = 80
        ws.column_dimensions["B"].width = 60
        
        # Emri i agjentit në krye (dinamik sipas gjuhës)
        agent_title = ws["A1"]
        if language == "en":
            agent_title.value = f"Agent: {agent}"
        elif language == "it":
            agent_title.value = f"Agente: {agent}"
        else:
            agent_title.value = f"Agjenti: {agent}"
        agent_title.font = Font(bold=True, size=16, color="1F4E79")
        agent_title.alignment = center
        
        # Strengths
        ws["A3"] = t["strengths"]
        ws["A3"].font = Font(bold=True, size=14, color="228B22")
        ws["A3"].fill = light_green_fill
        ws["A3"].border = thick_border
        
        # Lista e strengths me bullet points dhe renditje vertikale
        strengths_list = data["preggi"].split(" • ") if isinstance(data["preggi"], str) else data["preggi"]
        current_row = 4
        for i, strength in enumerate(strengths_list, 1):
            if strength.strip():
                cell = ws[f"A{current_row}"]
                cell.value = f"{i}. {strength.strip()}"
                cell.alignment = wrap
                cell.font = Font(size=12)
                cell.border = black_border
                current_row += 1
        
        # Improvements
        improvements_start = current_row + 2
        ws[f"A{improvements_start}"] = t["improvements"]
        ws[f"A{improvements_start}"].font = Font(bold=True, size=14, color="DC143C")
        ws[f"A{improvements_start}"].fill = light_blue_fill
        ws[f"A{improvements_start}"].border = thick_border
        
        # Lista e improvements me bullet points dhe renditje vertikale
        improvements_list = data["da_migliorare"].split(" • ") if isinstance(data["da_migliorare"], str) else data["da_migliorare"]
        current_row = improvements_start + 1
        for i, improvement in enumerate(improvements_list, 1):
            if improvement.strip():
                cell = ws[f"A{current_row}"]
                cell.value = f"{i}. {improvement.strip()}"
                cell.alignment = wrap
                cell.font = Font(size=12)
                cell.border = black_border
                current_row += 1
    
    # Ruaj rezultatin
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
