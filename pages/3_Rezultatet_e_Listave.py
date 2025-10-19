import io
import re
from datetime import datetime, time
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from core.db_vicidial import (
    fetch_outbound_by_list_statuses,
    get_inbound_calls_by_list,
    fetch_status_distribution_by_list,
    fetch_dials_by_phone,
    fetch_inbound_by_phone,
    fetch_svyclm_by_list,
    fetch_svyclm_timeout_by_list,
)
from core.voip_rates import get_voip_rates, update_voip_rates
from core.status_settings import (
    get_status_cost_map,
    get_resa_threshold_percent,
    get_dial_statuses_for_dials,
    get_min_dials_per_list,
    get_allow_all_statuses,
    update_dial_statuses_for_dials,
    update_allow_all_statuses,
)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from core.prefix_it import match_prefix, normalize_it_number

st.title("📈 Smart Report — VOIP Cost & Resa by List")

# Database selector
from core.db_vicidial import set_db_connection, get_current_db_key, _read_db_secrets
col_db1, col_db2 = st.columns([3, 1])
with col_db1:
    # Get host info for both databases
    db1_host, _, _, _ = _read_db_secrets("db")
    db2_host, _, _, _ = _read_db_secrets("db2")

    db_options = {
        f"Database 1 ({db1_host or 'Default'})": "db",
        f"Database 2 ({db2_host or '95.217.87.125'})": "db2"
    }
    db_label = st.selectbox(
        "Zgjidh Databazën",
        options=list(db_options.keys()),
        index=0 if get_current_db_key() == "db" else 1,
        help="Zgjidh nga cila databazë të lexohen të dhënat"
    )
    selected_db_key = db_options[db_label]
    set_db_connection(selected_db_key)
with col_db2:
    current_host, _, _, _ = _read_db_secrets(selected_db_key)
    st.caption(f"🔌 **{current_host}**")

st.markdown("---")

col1, col2 = st.columns(2)
with col1:
    start_date = st.date_input("Data fillimit")
    start_time = st.time_input("Ora fillimit", value=time(0,0,0))
with col2:
    end_date = st.date_input("Data mbarimit")
    end_time = st.time_input("Ora mbarimit", value=time(23,59,59))

campaign = st.text_input("Kampanja (p.sh. AUTOBIZ)", placeholder="Emri i kampanjës")
ivr_code = st.text_input("IVR code (p.sh. 1)", placeholder="vlera e butonit p.sh. 1")

# Status selector (before running)
st.markdown("### Dial statuses for 01_List_Cost")
allow_all_current = get_allow_all_statuses()
colS1, colS2 = st.columns([1,3])
with colS1:
    allow_all_new = st.checkbox("Allow ALL statuses", value=allow_all_current, help="Kur aktivizohet, nuk aplikohet asnjë filter statusi")
with colS2:
    current_statuses = get_dial_statuses_for_dials()
    # Lista e plotë e statuseve të mundshme në Vicidial
    all_possible_statuses = [
        "AA", "AB", "ADC", "AFAX", "AL", "AM", "A",
        "B", "BUSY",
        "CALLBK", "CBHOLD", "CPTERR",
        "DC", "DEC", "DISCONN", "DNC", "DROP", "DROPHG",
        "ERI", "ERRVM",
        "IVRXFR",
        "LAGGED", "LB", "LRERR",
        "MAXCALL", "MLINAT", "MLIVEO", "MLVMCT", "MLVMEO",
        "NA", "NANQUE", "NCALL", "NOINT", "N",
        "PAMD", "PDROP", "PM", "PU",
        "QVMAIL",
        "RQXFER",
        "SALE", "SVYCLM",
        "TIMEOT", "XDROP", "XFER"
    ]
    status_options = sorted(set(current_statuses + all_possible_statuses))
    selected_statuses = st.multiselect("Dial statuses (kur ALL është OFF)", options=status_options, default=current_statuses)
save_status = st.button("Ruaj statuset", use_container_width=False)
if save_status:
    try:
        update_allow_all_statuses(allow_all_new)
        update_dial_statuses_for_dials(selected_statuses)
        st.success("U ruajtën preferencat e statusit.")
    except Exception as e:
        st.error(f"Nuk u ruajtën: {e}")

col_type, col_mode = st.columns([2, 1])
with col_type:
    type_pref = st.radio("Type", ["all", "mobile", "fix"], index=0, horizontal=True)
with col_mode:
    show_full_report = st.checkbox("📊 Raport i Plotë", value=False, help="Shfaq tabelat e detajuara (01_List_Cost, 02_Prefix_Province, 03_SVYCLM_Quality)")

run = st.button("Gjenero raportin", type="primary", disabled=not (campaign and ivr_code))


def infer_list_type(list_name: str) -> str:
    n = (list_name or "").lower()
    if re.search(r"(mobile|cell|cellulare|gsm|mob)", n):
        return "mobile"
    if re.search(r"(fix|fisso|landline|fixed)", n):
        return "fix"
    return "unknown"


if run:
    from_ts = datetime.combine(start_date, start_time).strftime("%Y-%m-%d %H:%M:%S")
    to_ts = datetime.combine(end_date, end_time).strftime("%Y-%m-%d %H:%M:%S")

    # ============== AZHORNIMI I FILE-IT analysis_data_db.json ==============
    if show_full_report:
        prog = st.progress(0, text="🔄 Azhornohet analysis_data_db.json...")
        try:
            # Importojmë funksionet e nevojshme
            from core.db_vicidial import fetch_dials_by_phone, fetch_inbound_by_phone
            from core.mobile_fix_classifier import classify_phone_number
            import json
            from datetime import datetime

            # Merr të dhënat për Analyzer + Recommender
            prog.progress(10, text="Duke mbledhur të dhënat për Analyzer...")

            # Merr të dhënat e dials
            dial_statuses = None if get_allow_all_statuses() else get_dial_statuses_for_dials()
            if dial_statuses is not None and not dial_statuses:
                st.warning("No dial statuses selected. Shko te Settings për t'i vendosur ose aktivizo ALL.")
                st.stop()

            # Merr të dhënat e dials me provincë
            dials_data = fetch_dials_by_phone(from_ts, to_ts, campaign.strip(), dial_statuses)
            prog.progress(30, text="Duke mbledhur të dhënat e inbound...")

            # Merr të dhënat e inbound me provincë
            inbound_data = fetch_inbound_by_phone(from_ts, to_ts, campaign.strip(), ivr_code.strip())
            prog.progress(50, text="Duke analizuar të dhënat...")

            # Krijo strukturën e të dhënave për Analyzer
            analysis_data = {
                "collection_date": datetime.now().isoformat(),
                "campaign_id": campaign.strip(),
                "db_key": selected_db_key,  # Përdor database-in e zgjedhur
                "analysis_period_days": (datetime.strptime(to_ts, "%Y-%m-%d %H:%M:%S") - datetime.strptime(from_ts, "%Y-%m-%d %H:%M:%S")).days,
                "from_date": from_ts,
                "to_date": to_ts,
                "prefix_analysis": [],
                "status_distribution": [],
                "hourly_performance": [],
                "daily_performance": [],
                "list_performance": [],
                "recycling_status": [],
                "custom_fields": [],
                "active_lists": [],
                "closer_log": [],
                "sample_leads": [],
                "call_time_config": [],
                "lead_filter_config": [],
                "lead_filter_rules": [],
                "hopper_status": [],
                "prefix_status_analysis": [],
                "status_definitions": []
            }

            # Analizo të dhënat e dials
            phone_stats = {}
            for row in dials_data:
                phone = row.get("phone_number", "")
                province = row.get("province", "")
                calls = row.get("dials", 0)
                total_sec = row.get("total_sec", 0)

                if phone:
                    # Klasifikon numrin
                    phone_type, province_code, zone_name = classify_phone_number(phone, province)

                    if phone not in phone_stats:
                        phone_stats[phone] = {
                            "phone": phone,
                            "phone_type": phone_type,
                            "province": province_code,
                            "zone": zone_name,
                            "calls": 0,
                            "total_sec": 0,
                            "inbound_calls": 0
                        }

                    phone_stats[phone]["calls"] += calls
                    phone_stats[phone]["total_sec"] += total_sec

            # Analizo të dhënat e inbound
            for row in inbound_data:
                phone = row.get("phone_number", "")
                inbound_calls = row.get("inbound_calls", 0)

                if phone in phone_stats:
                    phone_stats[phone]["inbound_calls"] += inbound_calls

            # Krijo prefix_analysis
            prefix_stats = {}
            for phone, stats in phone_stats.items():
                if stats["phone_type"] == "FIX" and stats["province"]:
                    # Për numra fix, përdor provincën
                    province = stats["province"]
                    if province not in prefix_stats:
                        prefix_stats[province] = {
                            "prefix_2": "",
                            "prefix_3": "",
                            "prefix_4": "",
                            "calls": 0,
                            "avg_duration": 0,
                            "total_minutes": 0,
                            "inbound_calls": 0
                        }

                    prefix_stats[province]["calls"] += stats["calls"]
                    prefix_stats[province]["total_minutes"] += stats["total_sec"] / 60
                    prefix_stats[province]["inbound_calls"] += stats["inbound_calls"]

            # Konverto në listë
            analysis_data["prefix_analysis"] = [
                {
                    "prefix_2": "",
                    "prefix_3": "",
                    "prefix_4": "",
                    "calls": stats["calls"],
                    "avg_duration": stats["total_minutes"] / stats["calls"] if stats["calls"] > 0 else 0,
                    "total_minutes": stats["total_minutes"],
                    "inbound_calls": stats["inbound_calls"]
                }
                for province, stats in prefix_stats.items()
            ]

            # Ruaj file-in me emrin e duhur për database-in e zgjedhur
            suffix = selected_db_key.replace("/", "_")
            output_file = f"vicidial_analysis_data_{suffix}.json"
            prog.progress(80, text=f"Duke ruajtur {output_file}...")
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(analysis_data, f, indent=2, ensure_ascii=False, default=str)

            prog.progress(100, text=f"✅ {output_file} u azhornua!")

        except Exception as e:
            st.error(f"Gabim gjatë azhornimit të {output_file}: {e}")
            st.stop()

    # ============== LOGJIKA AKTUALE PËR RAPORTIN ==============
    prog = st.progress(0, text="Duke lexuar OUTBOUND...")
    try:
        dial_statuses = None if get_allow_all_statuses() else get_dial_statuses_for_dials()
        if dial_statuses is not None and not dial_statuses:
            st.warning("No dial statuses selected. Shko te Settings për t'i vendosur ose aktivizo ALL.")
            st.stop()
        ob_rows = fetch_outbound_by_list_statuses(from_ts, to_ts, campaign.strip(), dial_statuses)
        prog.progress(50, text="Duke lexuar INBOUND sipas IVR...")
        inbound_map = get_inbound_calls_by_list(from_ts, to_ts, campaign.strip(), ivr_code.strip())
        prog.progress(100, text="✅ Të dhënat u lexuan")
    except Exception as e:
        st.error(f"Gabim gjatë leximit të DB: {e}")
        st.stop()

    rates = get_voip_rates()
    status_costs = get_status_cost_map()
    resa_threshold = get_resa_threshold_percent()
    min_dials_per_list = get_min_dials_per_list()

    # -------------- Sheet 1: 01_List_Cost --------------
    records: List[Dict[str, Any]] = []
    total_dials_sum = 0
    inbound_calls_sum = 0
    total_voip_cost = 0.0

    for r in ob_rows or []:
        list_id = r.get("list_id")
        list_name = r.get("list_name")
        total_dials = int(r.get("total_dials") or 0)
        total_sec = float(r.get("total_sec") or 0)
        inbound_calls = int(inbound_map.get(int(list_id), 0))

        ltype = infer_list_type(list_name)
        total_min = total_sec / 60.0
        if ltype == "mobile":
            rate = rates.mobile_eur_per_min
        elif ltype == "fix":
            rate = rates.fix_eur_per_min
        else:
            rate = max(rates.mobile_eur_per_min, rates.fix_eur_per_min)

        voip_cost = total_min * rate if rate else None
        resa_pct = (inbound_calls / total_dials * 100.0) if total_dials else None
        cpi = (voip_cost / inbound_calls) if inbound_calls and voip_cost is not None else None

        if total_dials:
            total_dials_sum += total_dials
        if inbound_calls:
            inbound_calls_sum += inbound_calls
        if voip_cost is not None:
            total_voip_cost += voip_cost

        records.append({
            "list_id": list_id,
            "list_name": list_name,
            "list_type": ltype,
            "total_dials": total_dials,
            "inbound_calls": inbound_calls,
            "resa_%": round(resa_pct, 2) if resa_pct is not None else None,
            "total_min": round(total_min, 2),
            "voip_cost_eur": round(voip_cost, 4) if voip_cost is not None else None,
            "cost_per_inbound_eur": round(cpi, 4) if cpi is not None else None,
        })

    type_filter = type_pref

    records.sort(key=lambda x: (
        x["cost_per_inbound_eur"] if x["cost_per_inbound_eur"] is not None else 1e9,
        -(x["resa_%"] or 0)
    ))

    if type_filter != "all":
        records = [r for r in records if r["list_type"] == type_filter]

    # Recalculate totals after filtering
    total_dials_sum = sum(r.get("total_dials", 0) for r in records)
    inbound_calls_sum = sum(r.get("inbound_calls", 0) for r in records)
    total_voip_cost = sum(r.get("voip_cost_eur", 0) or 0 for r in records)
    total_minutes = sum(r.get("total_min", 0) for r in records)

    df = pd.DataFrame.from_records(records, columns=[
        "list_id","list_name","list_type","total_dials","inbound_calls","resa_%","total_min","voip_cost_eur","cost_per_inbound_eur"
    ])

    # smart_pick flag
    try:
        median_cpi = float(df["cost_per_inbound_eur"].dropna().median()) if not df.empty else None
    except Exception:
        median_cpi = None
    if median_cpi is not None:
        df["smart_pick"] = (
            (df["cost_per_inbound_eur"] <= median_cpi)
            & (df["resa_%"].fillna(0.0) >= resa_threshold)
            & (df["total_dials"].fillna(0) >= min_dials_per_list)
        )
    else:
        df["smart_pick"] = False

    # -------------- Summary Report (Visual) --------------
    st.markdown("### 📊 Përmbledhje e Raportit")

    # Display metrics in a nice layout
    col_kpi1, col_kpi2, col_kpi3, col_kpi4 = st.columns(4)

    with col_kpi1:
        st.metric(
            label="📞 Total Telefonata",
            value=f"{total_dials_sum:,}",
            help="Numri total i thirrjeve të bëra"
        )

    with col_kpi2:
        st.metric(
            label="📲 Inbound Calls",
            value=f"{inbound_calls_sum:,}",
            help="Thirrje që klienti ka shtypur kodin IVR"
        )

    with col_kpi3:
        st.metric(
            label="⏱️ Total Minuta",
            value=f"{total_minutes:,.0f}",
            help="Kohëzgjatja totale e thirrjeve në minuta"
        )

    with col_kpi4:
        st.metric(
            label="💰 Kosto VoIP",
            value=f"€ {total_voip_cost:,.2f}",
            help="Kostoja totale VoIP bazuar në tarifat e vendosura"
        )

    # Calculate Resa
    resa_overall = (inbound_calls_sum / total_dials_sum * 100.0) if total_dials_sum else 0.0

    # Additional info row
    col_info1, col_info2, col_info3, col_info4 = st.columns(4)
    with col_info1:
        st.caption(f"📅 Periudha: {from_ts[:10]} deri {to_ts[:10]}")
    with col_info2:
        st.caption(f"🎯 Resa: **{resa_overall:.2f}%**")
    with col_info3:
        avg_duration = (total_minutes / total_dials_sum) if total_dials_sum else 0
        st.caption(f"⏳ Kohëzgjatja mesatare: {avg_duration:.2f} min")
    with col_info4:
        cost_per_call = (total_voip_cost / total_dials_sum) if total_dials_sum else 0
        st.caption(f"💵 Kosto/thirrje: €{cost_per_call:.4f}")

    st.markdown("---")

    # Store summary in session state
    summary_data = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "campaign": campaign.strip(),
        "ivr_code": ivr_code.strip(),
        "from_date": from_ts[:10],
        "to_date": to_ts[:10],
        "total_dials": total_dials_sum,
        "inbound_calls": inbound_calls_sum,
        "total_minutes": round(total_minutes, 2),
        "voip_cost_eur": round(total_voip_cost, 2),
        "resa_percent": round(resa_overall, 2),
        "avg_duration_min": round(avg_duration, 2),
        "cost_per_call_eur": round(cost_per_call, 4),
    }
    st.session_state["last_smart_report"] = summary_data

    # If not full report, stop here
    if not show_full_report:
        st.success("✅ Përmbledhja u ruajt në **Raporte**. Për të eksportuar XLSX, aktivizo '📊 Raport i Plotë'.")
        st.stop()

    # -------------- FULL REPORT MODE --------------
    # Show detailed report only if checkbox is enabled
    st.markdown("### 📋 Raport i Plotë — Detaje")
    st.dataframe(df, use_container_width=True)

    # Now fetch additional data for full report
    prog_full = st.progress(0, text="Duke lexuar të dhëna shtesë për raportin e plotë...")
    try:
        dist_rows = fetch_status_distribution_by_list(from_ts, to_ts, campaign.strip())
        prog_full.progress(33, text="Duke lexuar prefix/provincia...")
        dials_phone = fetch_dials_by_phone(from_ts, to_ts, campaign.strip(), None)
        inbound_phone = fetch_inbound_by_phone(from_ts, to_ts, campaign.strip(), ivr_code.strip())
        prog_full.progress(66, text="Duke lexuar SVYCLM quality...")
        sv_rows = fetch_svyclm_by_list(from_ts, to_ts, campaign.strip())
        timeout_codes = ["TIMEOUT", "t", "TIME-OUT"]
        to_rows = fetch_svyclm_timeout_by_list(from_ts, to_ts, campaign.strip(), timeout_codes)
        prog_full.progress(100, text="✅ Të dhënat shtesë u lexuan")
    except Exception as e:
        st.error(f"Gabim gjatë leximit të të dhënave shtesë: {e}")
        st.stop()

    # -------------- Sheet 2: 02_Status_Mix_Cost --------------
    per_list: Dict[int, Dict[str, Any]] = {}
    for r in dist_rows or []:
        lid = int(r.get("list_id"))
        status = str(r.get("status") or "").upper()
        calls = int(r.get("calls") or 0)
        total_sec = float(r.get("total_sec") or 0)
        d = per_list.setdefault(lid, {"list_id": lid, "total_dials": 0, "total_sec": 0.0})
        d["total_dials"] += calls
        d["total_sec"] += total_sec
        d[f"{status}_calls"] = d.get(f"{status}_calls", 0) + calls

    name_map = {int(r.get("list_id")): r.get("list_name") for r in ob_rows}
    for lid, d in per_list.items():
        d["list_name"] = name_map.get(lid, f"LIST {lid}")
        d["inbound_calls"] = int(inbound_map.get(lid, 0))
        total_cost = 0.0
        for k, v in list(d.items()):
            if k.endswith("_calls"):
                status = k[:-6]
                cost = float(status_costs.get(status, 0.0))
                total_cost += float(v) * cost
        d["status_cost_total_eur"] = round(total_cost, 4)
        td = int(d.get("total_dials") or 0)
        ib = int(d.get("inbound_calls") or 0)
        d["status_cost_per_dial_eur"] = round(total_cost / td, 6) if td else None
        d["status_cost_per_inbound_eur"] = round(total_cost / ib, 6) if ib else None

    df_status = pd.DataFrame.from_records(list(per_list.values()))
    if not df_status.empty and "status_cost_per_inbound_eur" in df_status:
        df_status = df_status.sort_values(by=["status_cost_per_inbound_eur"], ascending=[True])

    # -------------- Sheet 2b: 02_Prefix_Province --------------
    # Build phone-level aggregation (data already fetched above)
    # LOGJIKA E RE: Përdor classify_phone_number() me provincën nga Vicidial
    from core.mobile_fix_classifier import classify_phone_number
    ib_map_phone = {str(r.get("phone_number")): int(r.get("inbound_calls") or 0) for r in inbound_phone or []}
    rows_prefix = []
    for r in dials_phone or []:
        phone = str(r.get("phone_number"))
        dials = int(r.get("dials") or 0)
        total_min = float(r.get("total_sec") or 0) / 60.0
        inbound_calls = int(ib_map_phone.get(phone, 0))

        # Merr provincën nga Vicidial (nëse ka)
        lead_province = r.get("province", "")

        # Përdor logjikën e re të klasifikimit
        phone_type, provincia, zone_name = classify_phone_number(phone, lead_province)

        # Përcakto rate-in bazuar në llojin e numrit
        if phone_type == "FIX":
            rate = rates.fix_eur_per_min
        elif phone_type == "MOBILE":
            rate = rates.mobile_eur_per_min
        else:
            rate = max(rates.mobile_eur_per_min, rates.fix_eur_per_min)

        voip_cost = total_min * rate
        resa_pct = (inbound_calls / dials * 100.0) if dials else None
        cpi = (voip_cost / inbound_calls) if inbound_calls else None
        rows_prefix.append({
            "provincia": provincia,
            "total_dials": dials,
            "inbound_calls": inbound_calls,
            "resa_%": round(resa_pct, 3) if resa_pct is not None else None,
            "total_min": round(total_min, 3),
            "voip_cost_eur": round(voip_cost, 3),
            "cost_per_inbound_eur": round(cpi, 3) if cpi is not None else None,
        })
    # Aggregate by prefix/provincia
    if rows_prefix:
        df_prefix = pd.DataFrame.from_records(rows_prefix)
        if not df_prefix.empty:
            agg_funcs = {
                "total_dials": "sum",
                "inbound_calls": "sum",
                "total_min": "sum",
                "voip_cost_eur": "sum",
            }
            df_prefix = df_prefix.groupby(["provincia"], as_index=False).agg(agg_funcs)
            df_prefix["resa_%"] = (df_prefix["inbound_calls"] / df_prefix["total_dials"] * 100.0).round(3)
            df_prefix["cost_per_inbound_eur"] = (df_prefix["voip_cost_eur"] / df_prefix["inbound_calls"]).where(df_prefix["inbound_calls"]>0).round(3)
            df_prefix = df_prefix.sort_values(by=["cost_per_inbound_eur","resa_%"], ascending=[True, False])
        else:
            df_prefix = pd.DataFrame(columns=["provincia","total_dials","inbound_calls","resa_%","total_min","voip_cost_eur","cost_per_inbound_eur"])
    else:
        df_prefix = pd.DataFrame(columns=["provincia","total_dials","inbound_calls","resa_%","total_min","voip_cost_eur","cost_per_inbound_eur"])

    # -------------- Merge List Cost with Status Mix --------------
    # Join df (list cost) with df_status on list_id (only if both have data)
    if not df.empty and not df_status.empty and "list_id" in df_status.columns:
        df_merged = df.merge(df_status, how="left", on=["list_id"], suffixes=("", "_status"))
        # Ensure list_name from left is kept; drop duplicated list_name_status if present
        if "list_name_status" in df_merged.columns:
            df_merged = df_merged.drop(columns=["list_name_status"], errors="ignore")
        # Drop duplicate total_dials from status side if present
        if "total_dials_status" in df_merged.columns:
            df_merged = df_merged.drop(columns=["total_dials_status"], errors=True)
        # Keep only PU_calls and SVYCLM_calls from any *_calls columns
        call_cols = [c for c in df_merged.columns if c.endswith("_calls")]
        cols_to_drop = [c for c in call_cols if c not in ("PU_calls", "SVYCLM_calls")]
        if cols_to_drop:
            df_merged = df_merged.drop(columns=cols_to_drop, errors=True)
        # Drop status_cost_* if all zeros or NaN
        for c in ["status_cost_total_eur","status_cost_per_dial_eur","status_cost_per_inbound_eur"]:
            if c in df_merged.columns:
                s = pd.to_numeric(df_merged[c], errors="coerce").fillna(0)
                if (s == 0).all():
                    df_merged = df_merged.drop(columns=[c], errors=True)
        # Round float columns to 3 decimals
        float_cols = df_merged.select_dtypes(include=["float"]).columns
        if len(float_cols) > 0:
            df_merged[float_cols] = df_merged[float_cols].round(3)
    else:
        # If no status data, use df as is
        df_merged = df.copy()
        # Round float columns to 3 decimals
        float_cols = df_merged.select_dtypes(include=["float"]).columns
        if len(float_cols) > 0:
            df_merged[float_cols] = df_merged[float_cols].round(3)

    # -------------- Sheet 4: 03_SVYCLM_Quality --------------
    # Data already fetched above
    from core.status_settings import get_svyclm_timeout_ratio_warn
    to_map = {int(r.get("list_id")): int(r.get("svyclm_timeout") or 0) for r in to_rows or []}
    qual_records = []
    warn_ratio = get_svyclm_timeout_ratio_warn()
    name_map2 = name_map
    for r in sv_rows or []:
        lid = int(r.get("list_id"))
        sv_calls = int(r.get("svyclm_calls") or 0)
        sv_timeout = int(to_map.get(lid, 0))
        li_total = next((x for x in records if x["list_id"] == lid), None)
        total_dials_l = int(li_total.get("total_dials") if li_total else 0)
        inbound_l = int(inbound_map.get(lid, 0))
        resa_l = (inbound_l / total_dials_l * 100.0) if total_dials_l else None
        ratio = (sv_timeout / sv_calls) if sv_calls else None
        note = "⚠ high timeout" if (ratio is not None and ratio >= warn_ratio) else ""
        qual_records.append({
            "list_id": lid,
            "list_name": name_map2.get(lid, f"LIST {lid}"),
            "total_dials": total_dials_l,
            "svyclm_calls": sv_calls,
            "svyclm_timeout": sv_timeout,
            "svyclm_timeout_ratio": round(ratio, 3) if ratio is not None else None,
            "inbound_calls": inbound_l,
            "resa_%": round(resa_l, 2) if resa_l is not None else None,
            "notes": note,
        })
    if qual_records:
        df_sv = pd.DataFrame.from_records(qual_records)
        # Round floats to 3 decimals
        fc_sv = df_sv.select_dtypes(include=["float"]).columns
        if len(fc_sv) > 0:
            df_sv[fc_sv] = df_sv[fc_sv].round(3)
    else:
        df_sv = pd.DataFrame(columns=["list_id","list_name","total_dials","svyclm_calls","svyclm_timeout","svyclm_timeout_ratio","inbound_calls","resa_%","notes"])

    # -------------- Sheet 3: 03_SVYCLM_Quality --------------

    # -------------- Export XLSX (3 sheets + metadata) --------------
    def export_xlsx_bytes(df_main: pd.DataFrame, df_prefix_only: pd.DataFrame, df_sv_only: pd.DataFrame) -> bytes:
        from io import BytesIO
        output = BytesIO()
        wb = Workbook()
        # 00_Run_Metadata
        ws0 = wb.active
        ws0.title = "00_Run_Metadata"
        ws0.append(["from_ts", from_ts])
        ws0.append(["to_ts", to_ts])
        ws0.append(["campaign", campaign.strip()])
        ws0.append(["ivr_code", ivr_code.strip()])
        status_mode = "ALL" if get_allow_all_statuses() else "FILTERED"
        ws0.append(["dial_statuses_mode", status_mode])
        if status_mode == "FILTERED":
            ws0.append(["dial_statuses_for_dials", ",".join(get_dial_statuses_for_dials())])
        # 01 (List Cost + Status Mix)
        ws1 = wb.create_sheet("01_List_Cost")
        for r in dataframe_to_rows(df_main, index=False, header=True):
            ws1.append(r)
        # 02 (Prefix_Province by provincia)
        ws2 = wb.create_sheet("02_Prefix_Province")
        for r in dataframe_to_rows(df_prefix_only, index=False, header=True):
            ws2.append(r)
        # 03 (SVYCLM Quality)
        ws3 = wb.create_sheet("03_SVYCLM_Quality")
        for r in dataframe_to_rows(df_sv_only, index=False, header=True):
            ws3.append(r)
        wb.save(output)
        return output.getvalue()

    # Build provincia-only dataframe from df_prefix
    if not df_prefix.empty:
        df_prov = df_prefix.groupby(["provincia"], as_index=False).agg({
            "total_dials": "sum",
            "inbound_calls": "sum",
            "total_min": "sum",
            "voip_cost_eur": "sum",
        })
        df_prov["resa_%"] = (df_prov["inbound_calls"] / df_prov["total_dials"] * 100.0).round(2)
        df_prov["cost_per_inbound_eur"] = (df_prov["voip_cost_eur"] / df_prov["inbound_calls"]).where(df_prov["inbound_calls"]>0)
        df_prov = df_prov.sort_values(by=["cost_per_inbound_eur","resa_%"], ascending=[True, False])
        df_prov = df_prov[["provincia","total_dials","inbound_calls","resa_%","total_min","voip_cost_eur","cost_per_inbound_eur"]]
    else:
        df_prov = pd.DataFrame(columns=["provincia","total_dials","inbound_calls","resa_%","total_min","voip_cost_eur","cost_per_inbound_eur"])

    xlsx_bytes = export_xlsx_bytes(df_merged, df_prov, df_sv)
    prog_full.progress(100, text="✅ Raporti i plotë gati")

    from datetime import datetime as dt
    st.download_button(
        "📥 Export XLSX (Raport i Plotë)",
        data=xlsx_bytes,
        file_name=f"smart_reports_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ================== Analyzer + Recommender (IVR Dial 700) ==================
st.markdown("---")
st.markdown("## 🧠 Analyzer + Recommender (IVR Dial 700)")
st.caption("Analizë e avancuar e listave dhe rekomandime për konfigurime në Vicidial. Kjo pjesë është vetëm vizuale dhe NUK ndryshon konfigurimet. Tarifat VoIP merren nga Settings.")

import os as _os
import json as _json
from core.list_analyzer import generate_report as _generate_report

_col_a1, _col_a2 = st.columns([2, 1])
with _col_a1:
    # Zgjidh file-in nga lista e disponueshme në projekt
    try:
        import glob as _glob
        _files = sorted(_glob.glob("vicidial_analysis_data_*.json"))
        if _os.path.exists("vicidial_analysis_data.json"):
            # përfshi dhe emrin legacy në fund të listës
            _files.append("vicidial_analysis_data.json")
    except Exception:
        _files = ["vicidial_analysis_data.json"] if _os.path.exists("vicidial_analysis_data.json") else []

    # Krijon opsione descriptive: "Database 1 (db) - vicidial_analysis_data_db.json"
    _options = []
    _file_map = {}
    for _f in _files:
        if _f == "vicidial_analysis_data.json":
            # Check what DB this legacy file corresponds to
            _legacy_db_key = None
            try:
                with open(_f, "r", encoding="utf-8") as _lf:
                    _legacy_data = _json.load(_lf)
                    _legacy_db_key = _legacy_data.get("db_key")
            except Exception:
                pass
            if _legacy_db_key == "db":
                _label = f"Database 1 (db) - {_f} [legacy]"
            elif _legacy_db_key == "db2":
                _label = f"Database 2 (db2) - {_f} [legacy]"
            else:
                _label = f"Legacy dataset - {_f}"
        else:
            # Extract db_key from filename
            _match = re.search(r'vicidial_analysis_data_(.+)\.json', _f)
            if _match:
                _db_key = _match.group(1)
                if _db_key == "db":
                    _label = f"Database 1 (db) - {_f}"
                elif _db_key == "db2":
                    _label = f"Database 2 (db2) - {_f}"
                else:
                    _label = f"Dataset {_db_key} - {_f}"
            else:
                _label = _f
        _options.append(_label)
        _file_map[_label] = _f

    if not _options:
        _options = ["(asnjë dataset i gjetur)"]
        _file_map["(asnjë dataset i gjetur)"] = None

    # Default selection: file që përputhet me DB e zgjedhur
    _default_label = f"Database 1 (db) - vicidial_analysis_data_db.json" if selected_db_key == "db" else f"Database 2 (db2) - vicidial_analysis_data_db2.json"
    _index = 0
    if _default_label in _options:
        _index = _options.index(_default_label)
    elif _options and _options[0] != "(asnjë dataset i gjetur)":
        _index = 0

    _selected_label = st.selectbox(
        "Zgjidh dataset-in për Analyzer (offline)",
        options=_options,
        index=_index if _options else 0,
        help="Zgjidh nga file-at e gjeneruar me collect_vicidial_data.py. Këshillë: përdor atë që përputhet me DB e zgjedhur sipër."
    )
    _data_path = _file_map.get(_selected_label, None)

    # Kategoritë e Analizës
    st.markdown("### 📋 Zgjidh Kategoritë e Analizës")
    st.caption("Mund të zgjedhësh një, dy, ose të treja kategoritë për të gjeneruar strategjitë")

    _col_cat1, _col_cat2, _col_cat3 = st.columns(3)
    with _col_cat1:
        _show_status_based = st.checkbox(
            "STATUS-BASED (A, B, C)",
            value=True,
            help="Strategji bazuar në lead recycling sipas statuseve (PU, BUSY, NOINT, etj.)"
        )
    with _col_cat2:
        _show_time_list = st.checkbox(
            "TIME AND LIST BASED (G, H, I)",
            value=True,
            help="Strategji bazuar në listat dhe oraret optimale (pa provinca, pa status)"
        )
    with _col_cat3:
        _show_time_place = st.checkbox(
            "TIME AND PLACE BASED (D, E, F)",
            value=True,
            help="Strategji bazuar në provincat dhe oraret optimale"
        )

    # Kontrollo që të paktën një kategori të jetë zgjedhur
    if not (_show_status_based or _show_time_list or _show_time_place):
        st.warning("⚠️ Ju lutem zgjedhni të paktën një kategori për të gjeneruar analizën!")

with _col_a2:
    _run_analyzer = st.button("🔎 Gjenero Analyzer", use_container_width=True)

if _run_analyzer:
    try:
        _report = _generate_report(_data_path)

        # Verifikim: a përputhet DB e file-it me zgjedhjen aktuale?
        _file_db = _report.get("campaign_id"), _report.get("analysis_period"), _report.get("mobile_vs_fix", {})
        _db_key_in_file = None
        try:
            raw = _json.load(open(_data_path, "r", encoding="utf-8"))
            _db_key_in_file = raw.get("db_key")
        except Exception:
            pass
        if _db_key_in_file and _db_key_in_file != selected_db_key:
            st.warning(f"Ky file është gjeneruar për DB '{_db_key_in_file}', ndërsa lart është zgjedhur '{selected_db_key}'.\nSugjerim: ekzekuto nga terminali → python collect_vicidial_data.py --db-key {selected_db_key} --campaign {campaign or 'autobiz'} dhe përdor file-in vicidial_analysis_data_{selected_db_key}.json")

        # Funnel
        _funnel = _report.get("press1_funnel", {}).get("funnel", {})
        _daily = _report.get("press1_funnel", {}).get("daily_metrics", {})

        st.markdown("### 📊 Funnel — Press 1")
        k1, k2, k3, k4 = st.columns(4)
        with k1:
            st.metric("Dials (7d)", f"{int(_funnel.get('total_dials_per_week', 0)):,}")
        with k2:
            st.metric("PU (7d)", f"{int(_funnel.get('pu_pick_up', 0)):,}")
        with k3:
            st.metric("SVYCLM (7d)", f"{int(_funnel.get('svyclm_listened_full', 0)):,}")
        with k4:
            st.metric("Press 1 (7d)", f"{int(_funnel.get('press1_actual', 0)):,}")

        k5, k6, k7, k8 = st.columns(4)
        with k5:
            st.metric("Dials→PU %", f"{_funnel.get('conversion_rates', {}).get('dials_to_pu', 0)}%")
        with k6:
            st.metric("Dials→SVYCLM %", f"{_funnel.get('conversion_rates', {}).get('dials_to_svyclm', 0)}%")
        with k7:
            st.metric("SVYCLM→Press1 %", f"{_funnel.get('conversion_rates', {}).get('svyclm_to_press1', 0)}%")
        with k8:
            st.metric("Press1 Overall %", f"{_funnel.get('conversion_rates', {}).get('overall_press1_rate', 0)}%")

        st.markdown("### 🎯 Objektivi ditor")
        g1, g2, g3 = st.columns(3)
        with g1:
            st.metric("Aktual", f"{_daily.get('current_press1_per_day', 0)} / ditë")
        with g2:
            st.metric("Target", f"{_daily.get('target_press1_per_day', 150)} / ditë")
        with g3:
            st.metric("Gap", f"{_daily.get('gap', 0)} ({_daily.get('improvement_needed_percent', 0)}%)")

        # Ranked lists
        _ranked = _report.get("ranked_lists", [])
        if _ranked:
            st.markdown("### 🏆 Top Lista")
            _df_rank = pd.DataFrame(_ranked)[[
                "list_id","list_name","phone_type","quality_score","available_leads","calls_7days","cost_per_call","recommendation"
            ]]
            st.dataframe(_df_rank.head(15), use_container_width=True)

        # Lead Recycling Analysis
        _recycling = _report.get("lead_recycling_analysis", {})
        if _recycling:
            st.markdown("### 🔄 Lead Recycling Analysis")
            st.caption("Analizë e statuseve për riciklimin e leads bazuar në performancë dhe kosto.")

            _recycle_summary = _recycling.get("summary", {})
            col_r1, col_r2, col_r3 = st.columns(3)
            with col_r1:
                st.metric("Total Calls", f"{_recycle_summary.get('total_calls_analyzed', 0):,}")
            with col_r2:
                st.metric("Recyclable %", f"{_recycle_summary.get('recyclable_percentage', 0):.1f}%")
            with col_r3:
                st.metric("Efficiency Gain", _recycle_summary.get("potential_efficiency_gain", "N/A"))

            # Status analysis table
            _status_data = []
            for category, statuses in _recycling.get("status_analysis", {}).items():
                for status in statuses:
                    _status_data.append({
                        "Status": status["status"],
                        "Calls": status["calls"],
                        "Percentage": f"{status['percentage']:.1f}%",
                        "Avg Duration": f"{status['avg_duration']:.1f}s",
                        "Category": category.replace("_", " ").title(),
                        "Recycle Attempts": status["attempt_maximum"],
                        "Delay Hours": status["attempt_delay"]
                    })

            if _status_data:
                st.markdown("#### 📊 Status Performance")
                st.dataframe(pd.DataFrame(_status_data), use_container_width=True)

        # Strategies (filtro bazuar në kategoritë e zgjedhura)
        _strats = _report.get("vicidial_recommendations", {}).get("strategies", [])

        # Filtro strategjitë bazuar në kategoritë e zgjedhura
        _filtered_strats = []
        for _strat in _strats:
            _strategy_id = _strat.get("strategy", "")
            # STATUS-BASED: A, B, C
            if _strategy_id in ["A", "B", "C"] and _show_status_based:
                _filtered_strats.append(_strat)
            # TIME AND LIST BASED: G, H, I
            elif _strategy_id in ["G", "H", "I"] and _show_time_list:
                _filtered_strats.append(_strat)
            # TIME AND PLACE BASED: D, E, F
            elif _strategy_id in ["D", "E", "F"] and _show_time_place:
                _filtered_strats.append(_strat)

        if _filtered_strats:
            st.markdown("### 🎯 Strategji për Vicidial")
            st.caption("Zgjidh strategjinë që përputhet me objektivat e tua dhe kopjo SQL scripts direkt në Vicidial.")

            # Shfaq kategoritë e zgjedhura
            _selected_categories = []
            if _show_status_based:
                _selected_categories.append("STATUS-BASED (A, B, C)")
            if _show_time_list:
                _selected_categories.append("TIME AND LIST BASED (G, H, I)")
            if _show_time_place:
                _selected_categories.append("TIME AND PLACE BASED (D, E, F)")

            st.info(f"📌 Kategoritë e zgjedhura: {', '.join(_selected_categories)}")

            for _strat in _filtered_strats:
                st.markdown(f"#### {_strat.get('strategy')}. {_strat.get('name')}")
                st.write(f"**Qëllimi:** {_strat.get('goal')}")
                st.write(f"**Fokusi:** {_strat.get('focus', 'N/A')}")

                _expected = _strat.get("expected_results", {})
                col_r1, col_r2, col_r3 = st.columns(3)
                with col_r1:
                    st.metric("Press1/ditë", _expected.get("press1_per_day", "N/A"))
                with col_r2:
                    st.metric("Ndikimi në kosto", _expected.get("cost_impact", "N/A"))
                with col_r3:
                    st.metric("Risk", _expected.get("risk", "N/A"))

                # STATUS-BASED: Tabela me vlera
                _recycle_table = _strat.get("recycle_table", [])
                if _recycle_table:
                    st.markdown("#### 📊 CAMPAIGN LEAD RECYCLE LISTINGS — Konfigurimi i Lead Recycling")
                    st.caption("Vendos manualisht këto vlera në Vicidial Admin → Campaign → Lead Recycle")

                    # Konverto në DataFrame për shfaqje më të mirë
                    import pandas as pd
                    df_recycle = pd.DataFrame(_recycle_table)

                    # Shfaq tabelën me kolonat kryesore
                    st.dataframe(
                        df_recycle[["STATUS", "ATTEMPT DELAY", "ATTEMPT MAXIMUM", "Arsyeja"]],
                        use_container_width=True,
                        hide_index=True
                    )

                    st.caption("💡 **Shpjegimi i kolonave:**")
                    st.caption("• **ATTEMPT DELAY**: Sa orë duhet të presë para se të rifutet në qarkullim një lead me këtë status")
                    st.caption("• **ATTEMPT MAXIMUM**: Sa herë maksimum duhet të tentohet një lead me këtë status para se të bllokohet")

                    if st.button(f"📋 Kopjo Tabela", key=f"copy_table_{_strat.get('strategy')}"):
                        st.success("✅ Tabela është kopjuar! Vendos manualisht në Vicidial.")

                # LIST-BASED: SQL Script (kushti pa WHERE dhe pa komente) + udhëzime jashtë
                _sql_script = _strat.get("sql_script", "")
                if _sql_script:
                    st.markdown("#### 📋 LEAD FILTER LISTINGS — Kusht për t'u vendosur (pa WHERE)")
                    st.code(_sql_script.strip(), language="sql")
                    st.caption("Vendose këtë kusht direkt në fushën e filter-it (pa WHERE, pa komente)")
                    with st.expander("ℹ️ Udhëzime"):
                        st.markdown("- Vendos kushtin në LEAD FILTER LISTINGS")
                        st.markdown("- Apliko te listat me performancë të lartë")
                        st.markdown("- Monitoro 1 javë dhe ajuste oraret/provincat sipas rezultateve")
                    if st.button(f"📋 Kopjo Kushtin", key=f"copy_{_strat.get('strategy')}"):
                        st.success("✅ Kushti u kopjua! Ngjite në Vicidial Admin → Lead Filters.")

                st.markdown("---")

        # Lead Conversion Rate by Hour (Grafiku)
        st.markdown("### 📈 Lead Conversion Rate sipas Orës (9:00-18:00)")
        st.caption("Analizë e performancës së thirrjeve sipas orës së ditës për periudhën e analizuar")

        try:
            # Merr të dhënat orare nga data
            _raw_data = _json.load(open(_data_path, "r", encoding="utf-8"))
            _hourly_data = _raw_data.get("hourly_performance", [])

            if _hourly_data:
                # Filtro për oraret 9-18
                _filtered_hours = [h for h in _hourly_data if 9 <= h.get("hour", 0) <= 18]

                if _filtered_hours:
                    # Krijo DataFrame
                    _df_hourly = pd.DataFrame(_filtered_hours)

                    # Llogarit conversion rate
                    if "conversion_rate" in _df_hourly.columns:
                        # Përdor conversion rate që vjen nga query (PU + SVYCLM / total_calls * 100)
                        _df_hourly["conversion_rate"] = _df_hourly["conversion_rate"].fillna(0)
                    elif "pu_count" in _df_hourly.columns and "svyclm_count" in _df_hourly.columns and "total_calls" in _df_hourly.columns:
                        # Nëse nuk ka conversion_rate por ka komponentët, llogarit manualisht
                        _df_hourly["conversion_rate"] = (
                            (_df_hourly["pu_count"] + _df_hourly["svyclm_count"]) / _df_hourly["total_calls"] * 100
                        ).fillna(0)
                    else:
                        # Fallback: përdor avg_duration si proxy (për datasets e vjetra)
                        _df_hourly["conversion_rate"] = (
                            _df_hourly.get("avg_duration", 0) / 10
                        ).fillna(0)

                    if "total_calls" in _df_hourly.columns:

                        # Krijo grafikun me plotly
                        fig = go.Figure()

                        # Shto line chart për conversion rate
                        fig.add_trace(go.Scatter(
                            x=_df_hourly["hour"],
                            y=_df_hourly["conversion_rate"],
                            mode='lines+markers',
                            name='Conversion Rate',
                            line=dict(color='#1f77b4', width=3),
                            marker=dict(size=8, color='#1f77b4'),
                            hovertemplate='<b>Ora %{x}:00</b><br>Conversion: %{y:.2f}%<extra></extra>'
                        ))

                        # Shto një vijë mesatare
                        _avg_conversion = _df_hourly["conversion_rate"].mean()
                        fig.add_hline(
                            y=_avg_conversion,
                            line_dash="dash",
                            line_color="red",
                            annotation_text=f"Mesatarja: {_avg_conversion:.2f}%",
                            annotation_position="right"
                        )

                        # Konfiguro layout
                        fig.update_layout(
                            title="Lead Conversion Rate sipas Orës",
                            xaxis_title="Ora e Ditës",
                            yaxis_title="Conversion Rate (%)",
                            xaxis=dict(
                                tickmode='linear',
                                tick0=9,
                                dtick=1,
                                tickformat='%H:00'
                            ),
                            yaxis=dict(
                                ticksuffix='%'
                            ),
                            hovermode='x unified',
                            template='plotly_white',
                            height=400
                        )

                        st.plotly_chart(fig, use_container_width=True)

                        # Shfaq statistika
                        col_stat1, col_stat2, col_stat3 = st.columns(3)
                        with col_stat1:
                            _best_hour = _df_hourly.loc[_df_hourly["conversion_rate"].idxmax(), "hour"]
                            st.metric("Ora më e mirë", f"{int(_best_hour)}:00", f"{_df_hourly['conversion_rate'].max():.2f}%")
                        with col_stat2:
                            _worst_hour = _df_hourly.loc[_df_hourly["conversion_rate"].idxmin(), "hour"]
                            st.metric("Ora më e dobët", f"{int(_worst_hour)}:00", f"{_df_hourly['conversion_rate'].min():.2f}%")
                        with col_stat3:
                            st.metric("Conversion mesatar", f"{_avg_conversion:.2f}%")

                        st.caption("💡 Përdor këto të dhëna për të optimizuar oraret e thirrjeve në LEAD FILTER LISTINGS")
                    else:
                        st.info("Të dhënat orare nuk përmbajnë informacion të mjaftueshëm për conversion rate.")
                else:
                    st.info("Nuk ka të dhëna për oraret 9-18.")
            else:
                st.info("Nuk ka të dhëna orare në dataset. Ekzekuto collect_vicidial_data.py për të mbledhur të dhëna.")
        except Exception as _e:
            st.warning(f"Nuk mund të gjenerohej grafiku: {_e}")

        st.markdown("---")

        # Scenarios
        _scenarios = _report.get("vicidial_recommendations", {}).get("scenarios", {})
        if _scenarios:
            st.markdown("### 📊 Skenare të Mundshme")
            st.caption("Analizë e rezultateve të mundshme bazuar në veprimet e ndryshme.")

            for _scenario_key, _scenario in _scenarios.items():
                with st.expander(f"🔮 {_scenario.get('name')}"):
                    st.write(f"**Përshkrimi:** {_scenario.get('description')}")

                    _predictions = _scenario.get("predictions", {})
                    col_s1, col_s2, col_s3, col_s4 = st.columns(4)
                    with col_s1:
                        st.metric("Press1/ditë", _predictions.get("press1_per_day", "N/A"))
                    with col_s2:
                        st.metric("Kosto/ditë", _predictions.get("cost_per_day", "N/A"))
                    with col_s3:
                        st.metric("Gap vs Target", _predictions.get("gap_vs_target", "N/A"))
                    with col_s4:
                        st.metric("ROI Impact", _predictions.get("roi_impact", "N/A"))

                    if "risks" in _scenario:
                        st.markdown("**Risqe:**")
                        for risk in _scenario["risks"]:
                            st.write(f"• {risk}")

                    if "implementation" in _scenario:
                        st.markdown("**Implementimi:**")
                        for impl in _scenario["implementation"]:
                            st.write(f"• {impl}")

                    st.write(f"**Timeline:** {_scenario.get('timeline', 'N/A')}")
        st.success("✅ Analyzer u gjenerua.")
    except Exception as _e:
        st.error(f"❌ Nuk u gjenerua Analyzer: {_e}")

    # (Analyzer + Recommender moved outside 'if run')
